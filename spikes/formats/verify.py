#!/usr/bin/env python3
"""Independent-reader checks for artifacts emitted by the Rust format spike."""
from __future__ import annotations

import csv
import json
import subprocess
import sys
from pathlib import Path

EXPECTED_COLUMNS = [
    "code_01", "response_text", "unicode_text",
    "very_long_variable_name_that_exceeds_stata_thirty_two_characters", "a-b", "a b",
]
EXPECTED_ROWS = 19


def check(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def main() -> None:
    directory = Path(sys.argv[1] if len(sys.argv) > 1 else "verify-output")
    csv_path = directory / "adversarial.csv"
    with csv_path.open(newline="", encoding="utf-8") as handle:
        records = list(csv.reader(handle))
    check(records[0] == EXPECTED_COLUMNS, "CSV column order changed")
    check(len(records) - 1 == EXPECTED_ROWS, "CSV row count changed")
    check(records[1][0] == "001" and records[7][0] == "", "CSV string/missing behavior changed")
    check(records[4][1] == "line one\nline two", "CSV newline escaping changed")

    import duckdb
    import pyarrow.parquet as pq

    table = pq.read_table(directory / "adversarial.parquet")
    check(table.column_names == EXPECTED_COLUMNS, "PyArrow column order changed")
    check(table.schema.field("code_01").type == "string", "Parquet code is not UTF-8 string")
    check(table.column("code_01")[0].as_py() == "001", "Parquet leading zero changed")
    check(duckdb.sql("SELECT count(*) FROM read_parquet(?)", params=[str(directory / "adversarial.parquet")]).fetchone()[0] == EXPECTED_ROWS, "DuckDB Parquet row count changed")

    import openpyxl

    workbook = openpyxl.load_workbook(directory / "adversarial.xlsx", read_only=True, data_only=True)
    check(workbook.sheetnames == ["Data 1", "Variables"], "Excel sheet structure changed")
    sheet = workbook["Data 1"]
    check([cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))] == EXPECTED_COLUMNS, "Excel variable-name row changed")
    check(sheet.cell(3, 1).value == "001", "Excel leading zero changed")
    check(sheet.cell(6, 2).value == "line one\nline two", "Excel newline changed")

    stata = directory / "adversarial.dta"
    if stata.exists():
        import pandas as pd
        frame = pd.read_stata(stata, preserve_dtypes=False, convert_categoricals=False)
        check(len(frame) == EXPECTED_ROWS, "pandas Stata row count changed")
        command = ["Rscript", "-e", "library(haven); x <- read_dta(commandArgs(TRUE)[1]); stopifnot(nrow(x) == 19)", str(stata)]
        subprocess.run(command, check=True)
    else:
        print("Stata not checked: writer is explicitly blocked.")
    print(json.dumps({"status": "ok", "rows": EXPECTED_ROWS, "directory": str(directory)}))


if __name__ == "__main__":
    main()
